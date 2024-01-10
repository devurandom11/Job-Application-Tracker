import openpyxl
from openpyxl import Workbook
import os
import argparse
import sys
from urllib.parse import urlparse
from dateutil import parser as date_parser


def check_for_spreadsheet(filename):
    expected_headers = [
        "Job Title",
        "Website",
        "Date Applied",
        "Company",
        "Location",
        "Job Role",
    ]

    try:
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active
        actual_headers = [cell.value for cell in sheet[1]]

        return actual_headers == expected_headers
    except Exception as e:
        print(f"Error reading file: {e}")
        return False


def generate_new_filename(base_path, ext, start_index=1):
    index = start_index
    while True:
        new_filename = f"{base_path}_{index}{ext}"
        if not os.path.exists(new_filename):
            return new_filename
        index += 1


def parse_arguments():
    parser = argparse.ArgumentParser(description="Job Application Tracker")
    parser.add_argument("-f", "--file", help="File to save job details to")
    return parser.parse_args()


def get_file_name(args):
    filename = args.file if args.file else "job_application_tracker.xlsx"
    if os.path.exists(filename) and not check_for_spreadsheet(filename):
        return handle_existing_file(filename)
    return filename


def handle_existing_file(filename):
    attempts = 0
    while True:
        print("File exists but does not contain correct headers.")
        user_choice = input(
            "Would you like to create a new file in the same directory? (y/n) "
        ).lower()

        if user_choice == "y":
            path, ext = os.path.splitext(filename)
            return generate_new_filename(path, ext)
        elif user_choice == "n":
            print("Exiting...")
            sys.exit()
        else:
            print("Invalid choice. Please enter 'y' for yes or 'n' for no.")
            attempts += 1
            if attempts >= 5:
                print("Maximum attempts reached. Exiting...")
                sys.exit()


def get_valid_website():
    while True:
        website = input("Enter website: ").strip()
        if not website.startswith("https://"):
            website = "https://" + website

        try:
            result = urlparse(website)
            if all([result.scheme, result.netloc]):
                return website
            else:
                print(
                    "Invalid website format. Please enter a valid URL starting with 'https://'."
                )
        except Exception as e:
            print(f"Invalid website format: {e}. Please try again.")


def get_valid_date():
    while True:
        date_input = input("Enter date applied: ")
        try:
            return date_parser.parse(date_input).strftime("%Y-%m-%d")
        except ValueError:
            print("Invalid date format. Please try again.")


def get_user_details():
    job_title = input("Enter job title: ")
    website = get_valid_website()
    date_applied = get_valid_date()
    company = input("Enter company name: ")
    location = input("Enter location: ")
    job_role = input("Enter job role: ")
    return [job_title, website, date_applied, company, location, job_role]


def save_to_excel(job_details, filename):
    if not os.path.exists(filename):
        wb = Workbook()
        sheet = wb.active
        sheet.append(
            ["Job Title", "Website", "Date Applied", "Company", "Location", "Job Role"]
        )
    else:
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active

    sheet.append(job_details)
    wb.save(filename)

    print(f"Job details saved successfully to {filename}")


def main():
    args = parse_arguments()
    filename = get_file_name(args)
    print("Job Application Tracker\n")

    job_details = get_user_details()
    save_to_excel(job_details, filename)


if __name__ == "__main__":
    main()
